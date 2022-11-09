import openpyxl
import requests
import re
from lxml import html
from markdownify import markdownify as md
from datetime import datetime
from common.scraper import get_scraper_by_url
import logging
import pytz
from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from user_messages import api as msg
import django_rq
from common.utils import GenerateDateUUIDMediaFilePath
import os
from books.models import BookReview, Book, BookMark, BookTag
from movies.models import MovieReview, Movie, MovieMark, MovieTag
from music.models import AlbumReview, Album, AlbumMark, AlbumTag
from games.models import GameReview, Game, GameMark, GameTag
from common.scraper import DoubanAlbumScraper, DoubanBookScraper, DoubanGameScraper, DoubanMovieScraper
from PIL import Image
from io import BytesIO
import filetype
from common.models import MarkStatusEnum


logger = logging.getLogger(__name__)
tz_sh = pytz.timezone('Asia/Shanghai')


def fetch_remote_image(url):
    try:
        print(f'fetching remote image {url}')
        raw_img = None
        ext = None
        if settings.SCRAPESTACK_KEY is not None:
            dl_url = f'http://api.scrapestack.com/scrape?access_key={settings.SCRAPESTACK_KEY}&url={url}'
        elif settings.SCRAPERAPI_KEY is not None:
            dl_url = f'http://api.scraperapi.com?api_key={settings.SCRAPERAPI_KEY}&url={url}'
        else:
            dl_url = url
        img_response = requests.get(dl_url, timeout=settings.SCRAPING_TIMEOUT)
        raw_img = img_response.content
        img = Image.open(BytesIO(raw_img))
        img.load()  # corrupted image will trigger exception
        content_type = img_response.headers.get('Content-Type')
        ext = filetype.get_type(mime=content_type.partition(';')[0].strip()).extension
        f = GenerateDateUUIDMediaFilePath(None, "x." + ext, settings.MARKDOWNX_MEDIA_PATH)
        file = settings.MEDIA_ROOT + f
        local_url = settings.MEDIA_URL + f
        os.makedirs(os.path.dirname(file), exist_ok=True)
        img.save(file)
        # print(f'remote image saved as {local_url}')
        return local_url
    except Exception:
        print(f'unable to fetch remote image {url}')
        return url


class DoubanImporter:
    total = 0
    processed = 0
    skipped = 0
    imported = 0
    failed = []
    user = None
    visibility = 0
    file = None

    def __init__(self, user, visibility):
        self.user = user
        self.visibility = visibility

    def update_user_import_status(self, status):
        self.user.preference.import_status['douban_pending'] = status
        self.user.preference.import_status['douban_file'] = self.file
        self.user.preference.import_status['douban_visibility'] = self.visibility
        self.user.preference.import_status['douban_total'] = self.total
        self.user.preference.import_status['douban_processed'] = self.processed
        self.user.preference.import_status['douban_skipped'] = self.skipped
        self.user.preference.import_status['douban_imported'] = self.imported
        self.user.preference.import_status['douban_failed'] = self.failed
        self.user.preference.save(update_fields=['import_status'])

    def import_from_file(self, uploaded_file):
        try:
            wb = openpyxl.open(uploaded_file, read_only=True, data_only=True, keep_links=False)
            wb.close()
            file = settings.MEDIA_ROOT + GenerateDateUUIDMediaFilePath(None, "x.xlsx", settings.SYNC_FILE_PATH_ROOT)
            os.makedirs(os.path.dirname(file), exist_ok=True)
            with open(file, 'wb') as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)
            self.file = file
            self.update_user_import_status(2)
            jid = f'Douban_{self.user.id}_{os.path.basename(self.file)}'
            django_rq.get_queue('doufen').enqueue(self.import_from_file_task, job_id=jid)
        except Exception:
            return False
        # self.import_from_file_task(file, user, visibility)
        return True

    mark_sheet_config = {
        '想读': [MarkStatusEnum.WISH, DoubanBookScraper, Book, BookMark, BookTag],
        '在读': [MarkStatusEnum.DO, DoubanBookScraper, Book, BookMark, BookTag],
        '读过': [MarkStatusEnum.COLLECT, DoubanBookScraper, Book, BookMark, BookTag],
        '想看': [MarkStatusEnum.WISH, DoubanMovieScraper, Movie, MovieMark, MovieTag],
        '在看': [MarkStatusEnum.DO, DoubanMovieScraper, Movie, MovieMark, MovieTag],
        '想看': [MarkStatusEnum.COLLECT, DoubanMovieScraper, Movie, MovieMark, MovieTag],
        '想听': [MarkStatusEnum.WISH, DoubanAlbumScraper, Album, AlbumMark, AlbumTag],
        '在听': [MarkStatusEnum.DO, DoubanAlbumScraper, Album, AlbumMark, AlbumTag],
        '听过': [MarkStatusEnum.COLLECT, DoubanAlbumScraper, Album, AlbumMark, AlbumTag],
        '想玩': [MarkStatusEnum.WISH, DoubanGameScraper, Game, GameMark, GameTag],
        '在玩': [MarkStatusEnum.DO, DoubanGameScraper, Game, GameMark, GameTag],
        '玩过': [MarkStatusEnum.COLLECT, DoubanGameScraper, Game, GameMark, GameTag],
    }
    review_sheet_config = {
        '书评': [DoubanBookScraper, Book, BookReview],
        '影评': [DoubanMovieScraper, Movie, MovieReview],
        '乐评': [DoubanAlbumScraper, Album, AlbumReview],
        '游戏评论&攻略': [DoubanGameScraper, Game, GameReview],
    }
    mark_data = {}
    review_data = {}
    entity_lookup = {}

    def load_sheets(self):
        f = open(self.file, 'rb')
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True, keep_links=False)
        for data, config in [(self.mark_data, self.mark_sheet_config), (self.review_data, self.review_sheet_config)]:
            for name in config:
                data[name] = []
                if name in wb:
                    print(f'{self.user} parsing {name}')
                    for row in wb[name].iter_rows(min_row=2, values_only=True):
                        cells = [cell for cell in row]
                        if len(cells) > 6:
                            data[name].append(cells)
        for sheet in self.mark_data.values():
            for cells in sheet:
                # entity_lookup["title|rating"] = [(url, time), ...]
                k = f'{cells[0]}|{cells[5]}'
                v = (cells[3], cells[4])
                if k in self.entity_lookup:
                    self.entity_lookup[k].append(v)
                else:
                    self.entity_lookup[k] = [v]
        self.total = sum(map(lambda a: len(a), self.review_data.values()))

    def guess_entity_url(self, title, rating, timestamp):
        k = f'{title}|{rating}'
        if k not in self.entity_lookup:
            return None
        v = self.entity_lookup[k]
        if len(v) > 1:
            v.sort(key=lambda c: abs(timestamp - (datetime.strptime(c[1], "%Y-%m-%d %H:%M:%S") if type(c[1])==str else c[1]).replace(tzinfo=tz_sh)))
        return v[0][0]
        # for sheet in self.mark_data.values():
        #     for cells in sheet:
        #         if cells[0] == title and cells[5] == rating:
        #             return cells[3]

    def import_from_file_task(self):
        print(f'{self.user} import start')
        msg.info(self.user, f'开始导入豆瓣评论')
        self.update_user_import_status(1)
        self.load_sheets()
        print(f'{self.user} sheet loaded, {self.total} lines total')
        self.update_user_import_status(1)
        for name, param in self.review_sheet_config.items():
            self.import_review_sheet(self.review_data[name], param[0], param[1], param[2])
        self.update_user_import_status(0)
        msg.success(self.user, f'豆瓣评论导入完成，共处理{self.total}篇，已存在{self.skipped}篇，新增{self.imported}篇。')
        if len(self.failed):
            msg.error(self.user, f'豆瓣评论导入时未能处理以下网址：\n{" , ".join(self.failed)}')

    def import_review_sheet(self, worksheet, scraper, entity_class, review_class):
        prefix = f'{self.user} |'
        if worksheet is None:  # or worksheet.max_row < 2:
            print(f'{prefix} {review_class.__name__} empty sheet')
            return
        for cells in worksheet:
            if len(cells) < 6:
                continue
            title = cells[0]
            entity_title = re.sub('^《', '', re.sub('》$', '', cells[1]))
            review_url = cells[2]
            time = cells[3]
            rating = cells[4]
            content = cells[6]
            self.processed += 1
            if time:
                if type(time) == str:
                    time = datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
                time = time.replace(tzinfo=tz_sh)
            else:
                time = None
            if not content:
                content = ""
            if not title:
                title = ""
            r = self.import_review(entity_title, rating, title, review_url, content, time, scraper, entity_class, review_class)
            if r == 1:
                self.imported += 1
            elif r == 2:
                self.skipped += 1
            else:
                self.failed.append(review_url)
            self.update_user_import_status(1)

    def import_review(self, entity_title, rating, title, review_url, content, time, scraper, entity_class, review_class):
        # return 1: done / 2: skipped / None: failed
        prefix = f'{self.user} |'
        url = self.guess_entity_url(entity_title, rating, time)
        if url is None:
            print(f'{prefix} fetching {review_url}')
            try:
                if settings.SCRAPESTACK_KEY is not None:
                    _review_url = f'http://api.scrapestack.com/scrape?access_key={settings.SCRAPESTACK_KEY}&url={review_url}'
                else:
                    _review_url = review_url
                r = requests.get(_review_url, timeout=settings.SCRAPING_TIMEOUT)
                if r.status_code != 200:
                    print(f'{prefix} fetching error {review_url} {r.status_code}')
                    return
                h = html.fromstring(r.content.decode('utf-8'))
                for u in h.xpath("//header[@class='main-hd']/a/@href"):
                    if '.douban.com/subject/' in u:
                        url = u
                if not url:
                    print(f'{prefix} fetching error {review_url} unable to locate entity url')
                    return
            except Exception:
                print(f'{prefix} fetching exception {review_url}')
                return
        try:
            entity = entity_class.objects.get(source_url=url)
            print(f'{prefix} matched {url}')
        except ObjectDoesNotExist:
            try:
                print(f'{prefix} scraping {url}')
                scraper.scrape(url)
                form = scraper.save(request_user=self.user)
                entity = form.instance
            except Exception as e:
                print(f"{prefix} scrape failed: {url} {e}")
                logger.error(f"{prefix} scrape failed: {url}", exc_info=e)
                return
        params = {
            'owner': self.user,
            entity_class.__name__.lower(): entity
        }
        if review_class.objects.filter(**params).exists():
            return 2
        content = re.sub(r'<span style="font-weight: bold;">([^<]+)</span>', r'<b>\1</b>', content)
        content = re.sub(r'(<img [^>]+>)', r'\1<br>', content)
        content = re.sub(r'<div class="image-caption">([^<]+)</div>', r'<br><i>\1</i><br>', content)
        content = md(content)
        content = re.sub(r'(?<=!\[\]\()([^)]+)(?=\))', lambda x: fetch_remote_image(x[1]), content)
        params = {
            'owner': self.user,
            'created_time': time,
            'edited_time': time,
            'title': title,
            'content': content,
            'visibility': self.visibility,
            entity_class.__name__.lower(): entity,
        }
        review_class.objects.create(**params)
        return 1
