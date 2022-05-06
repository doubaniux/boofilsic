import logging
import pytz
from dataclasses import dataclass
from datetime import datetime
from django.conf import settings
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import load_workbook
from books.models import BookMark, Book, BookTag
from movies.models import MovieMark, Movie, MovieTag
from music.models import AlbumMark, Album, AlbumTag
from games.models import GameMark, Game, GameTag
from common.scraper import DoubanAlbumScraper, DoubanBookScraper, DoubanGameScraper, DoubanMovieScraper
from common.models import MarkStatusEnum
from .models import SyncTask


logger = logging.getLogger(__name__)


def __import_should_stop():
    # TODO: using queue.connection.set(job.key + b':should_stop', 1, ex=30) on the caller side and connection.get(job.key + b':should_stop') on the worker side.
    pass


def import_doufen_task(synctask):
    sync_doufen_job(synctask, __import_should_stop)


class DoufenParser:

    # indices in xlsx
    URL_INDEX = 4
    CONTENT_INDEX = 8
    TAG_INDEX = 7
    TIME_INDEX = 5
    RATING_INDEX = 6

    def __init__(self, task):
        self.__file_path = task.file.path
        self.__progress_sheet, self.__progress_row = task.get_breakpoint()
        self.__is_new_task = True
        if self.__progress_sheet is not None:
            self.__is_new_task = False
        if self.__progress_row is None:
            self.__progress_row = 2
        # data in the excel parse in python types
        self.task = task
        self.items = []

    def __open_file(self):
        self.__fp = open(self.__file_path, 'rb')
        self.__wb = load_workbook(
            self.__fp,
            read_only=True,
            data_only=True,
            keep_links=False
        )

    def __close_file(self):
        if self.__wb is not None:
            self.__wb.close()
        self.__fp.close()

    def __get_item_classes_mapping(self):
        '''
        We assume that the sheets names won't change
        '''
        mappings = []
        if self.task.sync_movie:
            for sheet_name in ['想看', '在看', '看过']:
                mappings.append({'sheet': sheet_name, 'mark_class': MovieMark,
                                 'entity_class': Movie, 'tag_class': MovieTag, 'scraper': DoubanMovieScraper})
        if self.task.sync_music:
            for sheet_name in ['想听', '在听', '听过']:
                mappings.append({'sheet': sheet_name, 'mark_class': AlbumMark,
                                 'entity_class': Album, 'tag_class': AlbumTag, 'scraper': DoubanAlbumScraper})
        if self.task.sync_book:
            for sheet_name in ['想读', '在读', '读过']:
                mappings.append({'sheet': sheet_name, 'mark_class': BookMark,
                                 'entity_class': Book, 'tag_class': BookTag, 'scraper': DoubanBookScraper})
        if self.task.sync_game:
            for sheet_name in ['想玩', '在玩', '玩过']:
                mappings.append({'sheet': sheet_name, 'mark_class': GameMark,
                                 'entity_class': Game, 'tag_class': GameTag, 'scraper': DoubanGameScraper})

        mappings.sort(key=lambda mapping: mapping['sheet'])

        if not self.__is_new_task:
            start_index = [mapping['sheet']
                           for mapping in mappings].index(self.__progress_sheet)
            mappings = mappings[start_index:]

        self.__mappings = mappings
        return mappings

    def __parse_items(self):
        assert self.__wb is not None, 'workbook not found'

        item_classes_mappings = self.__get_item_classes_mapping()

        is_first_sheet = True
        for mapping in item_classes_mappings:
            if mapping['sheet'] not in self.__wb:
                print(f"Sheet not found: {mapping['sheet']}")
                continue
            ws = self.__wb[mapping['sheet']]

            max_row = ws.max_row
            # empty sheet
            if max_row <= 1:
                continue

            # decide starting position
            start_row_index = 2
            if not self.__is_new_task and is_first_sheet:
                start_row_index = self.__progress_row

            # parse data
            i = start_row_index
            for row in ws.iter_rows(min_row=start_row_index, max_row=max_row, values_only=True):
                cells = [cell for cell in row]
                url = cells[self.URL_INDEX - 1]
                tags = cells[self.TAG_INDEX - 1]
                tags = list(set(tags.split(','))) if tags else None
                time = cells[self.TIME_INDEX - 1]
                if time:
                    time = datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
                    tz = pytz.timezone('Asia/Shanghai')
                    time = time.replace(tzinfo=tz)
                else:
                    time = None
                content = cells[self.CONTENT_INDEX - 1]
                if not content:
                    content = ""
                rating = cells[self.RATING_INDEX - 1]
                rating = int(rating) * 2 if rating else None
                self.items.append({
                    'data': DoufenRowData(url, tags, time, content, rating),
                    'entity_class': mapping['entity_class'],
                    'mark_class': mapping['mark_class'],
                    'tag_class': mapping['tag_class'],
                    'scraper': mapping['scraper'],
                    'sheet': mapping['sheet'],
                    'row_index': i,
                })
                i = i + 1

            # set first sheet flag
            is_first_sheet = False

    def __get_item_number(self):
        assert not self.__wb is None, 'workbook not found'
        assert not self.__mappings is None, 'mappings not found'

        sheets = [mapping['sheet'] for mapping in self.__mappings]
        item_number = 0
        for sheet in sheets:
            if sheet in self.__wb:
                item_number += self.__wb[sheet].max_row - 1

        return item_number

    def __update_total_items(self):
        total = self.__get_item_number()
        self.task.total_items = total
        self.task.save(update_fields=["total_items"])

    def parse(self):
        try:
            self.__open_file()
            self.__parse_items()
            if self.__is_new_task:
                self.__update_total_items()
            self.__close_file()
            return self.items
        except Exception as e:
            logger.error(f'Error parsing {self.__file_path} {e}')
            self.task.is_failed = True
        finally:
            self.__close_file()
        return []


@dataclass
class DoufenRowData:
    url: str
    tags: list
    time: datetime
    content: str
    rating: int


def add_new_mark(data, user, entity, entity_class, mark_class, tag_class, sheet, default_public):
    params = {
        'owner': user,
        'created_time': data.time,
        'edited_time': data.time,
        'rating': data.rating,
        'text': data.content,
        'status': translate_status(sheet),
        'visibility': 0 if default_public else 1,
        entity_class.__name__.lower(): entity,
    }
    mark = mark_class.objects.create(**params)
    entity.update_rating(None, data.rating)
    if data.tags:
        for tag in data.tags:
            params = {
                'content': tag,
                entity_class.__name__.lower(): entity,
                'mark': mark
            }
            try:
                tag_class.objects.create(**params)
            except Exception as e:
                logger.error(f'Error creating tag {tag} {mark}: {e}')


def overwrite_mark(entity, entity_class, mark, mark_class, tag_class, data, sheet):
    old_rating = mark.rating
    old_tags = getattr(mark, mark_class.__name__.lower() + '_tags').all()
    # update mark logic
    mark.created_time = data.time
    mark.edited_time = data.time
    mark.text = data.content
    mark.rating = data.rating
    mark.status = translate_status(sheet)
    mark.save()
    entity.update_rating(old_rating, data.rating)
    if old_tags:
        for tag in old_tags:
            tag.delete()
    if data.tags:
        for tag in data.tags:
            params = {
                'content': tag,
                entity_class.__name__.lower(): entity,
                'mark': mark
            }
            try:
                tag_class.objects.create(**params)
            except Exception as e:
                logger.error(f'Error creating tag {tag} {mark}: {e}')


def sync_doufen_job(task, stop_check_func):
    """
    TODO: Update task status every certain amount of items to reduce IO consumption
    """
    task = SyncTask.objects.get(pk=task.pk)
    if task.is_finished:
        return

    print(f'Task {task.pk}: loading')
    parser = DoufenParser(task)
    items = parser.parse()

    # use pop to reduce memo consumption
    while len(items) > 0 and not stop_check_func():
        item = items.pop(0)
        data = item['data']
        entity_class = item['entity_class']
        mark_class = item['mark_class']
        tag_class = item['tag_class']
        scraper = item['scraper']
        sheet = item['sheet']
        row_index = item['row_index']

        # update progress
        task.set_breakpoint(sheet, row_index, save=True)

        # scrape the entity if not exists
        try:
            entity = entity_class.objects.get(source_url=data.url)
            print(f'Task {task.pk}: {len(items)+1} remaining; matched {data.url}')
        except ObjectDoesNotExist:
            try:
                print(f'Task {task.pk}: {len(items)+1} remaining; scraping {data.url}')
                scraper.scrape(data.url)
                form = scraper.save(request_user=task.user)
                entity = form.instance
            except Exception as e:
                logger.error(f"Task {task.pk}: scrape failed: {data.url} {e}")
                if settings.DEBUG:
                    logger.error("Expections during scraping data:", exc_info=e)
                task.failed_urls.append(data.url)
                task.finished_items += 1
                task.save(update_fields=['failed_urls', 'finished_items'])
                continue

        # sync mark
        try:
            # already exists
            params = {
                'owner': task.user,
                entity_class.__name__.lower(): entity
            }
            mark = mark_class.objects.get(**params)

            if task.overwrite:
                overwrite_mark(entity, entity_class, mark,
                               mark_class, tag_class, data, sheet)
            else:
                task.success_items += 1
                task.finished_items += 1
                task.save(update_fields=['success_items', 'finished_items'])
                continue

        except ObjectDoesNotExist:
            add_new_mark(data, task.user, entity, entity_class,
                         mark_class, tag_class, sheet, task.default_public)

        except Exception as e:
            logger.error(
                f"Task {task.pk}: error when syncing marks", exc_info=e)
            task.failed_urls.append(data.url)
            task.finished_items += 1
            task.save(update_fields=['failed_urls', 'finished_items'])
            continue

        task.success_items += 1
        task.finished_items += 1
        task.save(update_fields=['success_items', 'finished_items'])

    # if task finish
    print(f'Task {task.pk}: stopping')
    if len(items) == 0:
        task.is_finished = True
        task.clear_breakpoint()
        task.save(update_fields=['is_finished', 'break_point'])


def translate_status(sheet_name):
    if '想' in sheet_name:
        return MarkStatusEnum.WISH
    elif '在' in sheet_name:
        return MarkStatusEnum.DO
    elif '过' in sheet_name:
        return MarkStatusEnum.COLLECT

    raise ValueError("Not valid status")
