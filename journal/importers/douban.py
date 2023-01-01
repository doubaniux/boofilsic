import openpyxl
import re
from markdownify import markdownify as md
from datetime import datetime
import logging
import pytz
from django.conf import settings
from user_messages import api as msg
import django_rq
from common.utils import GenerateDateUUIDMediaFilePath
import os
from catalog.common import *
from catalog.common.downloaders import *
from catalog.sites.douban import DoubanDownloader
from journal.models import *

_logger = logging.getLogger(__name__)
_tz_sh = pytz.timezone("Asia/Shanghai")


def _fetch_remote_image(url):
    try:
        print(f"fetching remote image {url}")
        imgdl = ProxiedImageDownloader(url)
        raw_img = imgdl.download().content
        ext = imgdl.extention
        f = GenerateDateUUIDMediaFilePath(
            None, "x." + ext, settings.MARKDOWNX_MEDIA_PATH
        )
        file = settings.MEDIA_ROOT + f
        local_url = settings.MEDIA_URL + f
        os.makedirs(os.path.dirname(file), exist_ok=True)
        with open(file, "wb") as binary_file:
            binary_file.write(raw_img)
        # print(f'remote image saved as {local_url}')
        return local_url
    except Exception:
        print(f"unable to fetch remote image {url}")
        return url


class DoubanImporter:
    total = 0
    processed = 0
    skipped = 0
    imported = 0
    failed = []
    user = None
    visibility = 0
    file = None

    def __init__(self, user, visibility):
        self.user = user
        self.visibility = visibility

    def update_user_import_status(self, status):
        self.user.preference.import_status["douban_pending"] = status
        self.user.preference.import_status["douban_file"] = self.file
        self.user.preference.import_status["douban_visibility"] = self.visibility
        self.user.preference.import_status["douban_total"] = self.total
        self.user.preference.import_status["douban_processed"] = self.processed
        self.user.preference.import_status["douban_skipped"] = self.skipped
        self.user.preference.import_status["douban_imported"] = self.imported
        self.user.preference.import_status["douban_failed"] = self.failed
        self.user.preference.save(update_fields=["import_status"])

    def import_from_file(self, uploaded_file):
        try:
            wb = openpyxl.open(
                uploaded_file, read_only=True, data_only=True, keep_links=False
            )
            wb.close()
            file = settings.MEDIA_ROOT + GenerateDateUUIDMediaFilePath(
                None, "x.xlsx", settings.SYNC_FILE_PATH_ROOT
            )
            os.makedirs(os.path.dirname(file), exist_ok=True)
            with open(file, "wb") as destination:
                for chunk in uploaded_file.chunks():
                    destination.write(chunk)
            self.file = file
            self.update_user_import_status(2)
            jid = f"Douban_{self.user.id}_{os.path.basename(self.file)}"
            django_rq.get_queue("import").enqueue(
                self.import_from_file_task, job_id=jid
            )
        except Exception:
            return False
        # self.import_from_file_task(file, user, visibility)
        return True

    mark_sheet_config = {
        "想读": [ShelfType.WISHLIST],
        "在读": [ShelfType.PROGRESS],
        "读过": [ShelfType.COMPLETE],
        "想看": [ShelfType.WISHLIST],
        "在看": [ShelfType.PROGRESS],
        "想看": [ShelfType.COMPLETE],
        "想听": [ShelfType.WISHLIST],
        "在听": [ShelfType.PROGRESS],
        "听过": [ShelfType.COMPLETE],
        "想玩": [ShelfType.WISHLIST],
        "在玩": [ShelfType.PROGRESS],
        "玩过": [ShelfType.COMPLETE],
    }
    review_sheet_config = {
        "书评": [Edition],
        "影评": [Movie],
        "乐评": [Album],
        "游戏评论&攻略": [Game],
    }
    mark_data = {}
    review_data = {}
    entity_lookup = {}

    def load_sheets(self):
        """Load data into mark_data / review_data / entity_lookup"""
        f = open(self.file, "rb")
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True, keep_links=False)
        for data, config in [
            (self.mark_data, self.mark_sheet_config),
            (self.review_data, self.review_sheet_config),
        ]:
            for name in config:
                data[name] = []
                if name in wb:
                    print(f"{self.user} parsing {name}")
                    for row in wb[name].iter_rows(min_row=2, values_only=True):
                        cells = [cell for cell in row]
                        if len(cells) > 6 and cells[0]:
                            data[name].append(cells)
        for sheet in self.mark_data.values():
            for cells in sheet:
                # entity_lookup["title|rating"] = [(url, time), ...]
                k = f"{cells[0]}|{cells[5]}"
                v = (cells[3], cells[4])
                if k in self.entity_lookup:
                    self.entity_lookup[k].append(v)
                else:
                    self.entity_lookup[k] = [v]
        self.total = sum(map(lambda a: len(a), self.mark_data.values()))
        self.total += sum(map(lambda a: len(a), self.review_data.values()))

    def guess_entity_url(self, title, rating, timestamp):
        k = f"{title}|{rating}"
        if k not in self.entity_lookup:
            return None
        v = self.entity_lookup[k]
        if len(v) > 1:
            v.sort(
                key=lambda c: abs(
                    timestamp
                    - (
                        datetime.strptime(c[1], "%Y-%m-%d %H:%M:%S")
                        if type(c[1]) == str
                        else c[1]
                    ).replace(tzinfo=_tz_sh)
                )
            )
        return v[0][0]
        # for sheet in self.mark_data.values():
        #     for cells in sheet:
        #         if cells[0] == title and cells[5] == rating:
        #             return cells[3]

    def import_from_file_task(self):
        print(f"{self.user} import start")
        msg.info(self.user, f"开始导入豆瓣标记和评论")
        self.update_user_import_status(1)
        self.load_sheets()
        print(f"{self.user} sheet loaded, {self.total} lines total")
        self.update_user_import_status(1)
        for name, param in self.mark_sheet_config.items():
            self.import_mark_sheet(self.mark_data[name], param[0], name)
        for name, param in self.review_sheet_config.items():
            self.import_review_sheet(self.review_data[name], name)
        self.update_user_import_status(0)
        msg.success(
            self.user,
            f"豆瓣标记和评论导入完成，共处理{self.total}篇，已存在{self.skipped}篇，新增{self.imported}篇。",
        )
        if len(self.failed):
            msg.error(self.user, f'豆瓣评论导入时未能处理以下网址：\n{" , ".join(self.failed)}')

    def import_mark_sheet(self, worksheet, shelf_type, sheet_name):
        prefix = f"{self.user} {sheet_name}|"
        if worksheet is None:  # or worksheet.max_row < 2:
            print(f"{prefix} empty sheet")
            return
        for cells in worksheet:
            if len(cells) < 6:
                continue
            # title = cells[0] or ""
            url = cells[3]
            time = cells[4]
            rating = cells[5]
            rating_grade = int(rating) * 2 if rating else None
            tags = cells[6] if len(cells) >= 7 else ""
            tags = tags.split(",") if tags else []
            comment = cells[7] if len(cells) >= 8 else None
            self.processed += 1
            if time:
                if type(time) == str:
                    time = datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
                time = time.replace(tzinfo=_tz_sh)
            else:
                time = None
            r = self.import_mark(url, shelf_type, comment, rating_grade, tags, time)
            if r == 1:
                self.imported += 1
            elif r == 2:
                self.skipped += 1
            self.update_user_import_status(1)

    def import_mark(self, url, shelf_type, comment, rating_grade, tags, time):
        """
        Import one mark: return 1: done / 2: skipped / None: failed
        """
        item = self.get_item_by_url(url)
        if not item:
            print(f"{self.user} | match/fetch {url} failed")
            return
        mark = Mark(self.user, item)
        if (
            mark.shelf_type == shelf_type
            or mark.shelf_type == ShelfType.COMPLETE
            or (
                mark.shelf_type == ShelfType.PROGRESS
                and shelf_type == ShelfType.WISHLIST
            )
        ):
            return 2
        mark.update(
            shelf_type, comment, rating_grade, self.visibility, created_time=time
        )
        if tags:
            TagManager.tag_item_by_user(item, self.user, tags)
        return 1

    def import_review_sheet(self, worksheet, sheet_name):
        prefix = f"{self.user} {sheet_name}|"
        if worksheet is None:  # or worksheet.max_row < 2:
            print(f"{prefix} empty sheet")
            return
        for cells in worksheet:
            if len(cells) < 6:
                continue
            title = cells[0]
            entity_title = (
                re.sub("^《", "", re.sub("》$", "", cells[1])) if cells[1] else ""
            )
            review_url = cells[2]
            time = cells[3]
            rating = cells[4]
            content = cells[6]
            self.processed += 1
            if time:
                if type(time) == str:
                    time = datetime.strptime(time, "%Y-%m-%d %H:%M:%S")
                time = time.replace(tzinfo=_tz_sh)
            else:
                time = None
            if not content:
                content = ""
            if not title:
                title = ""
            r = self.import_review(
                entity_title, rating, title, review_url, content, time
            )
            if r == 1:
                self.imported += 1
            elif r == 2:
                self.skipped += 1
            else:
                self.failed.append(review_url)
            self.update_user_import_status(1)

    def get_item_by_url(self, url):
        item = None
        try:
            site = SiteManager.get_site_by_url(url)
            item = site.get_item()
            if not item:
                print(f"fetching {url}")
                site.get_resource_ready()
                item = site.get_item()
                item.last_editor = user
                item.save()
            else:
                print(f"matched {url}")
        except Exception as e:
            print(f"fetching exception: {url} {e}")
            _logger.error(f"scrape failed: {url}", exc_info=e)
        if item is None:
            self.failed.append(url)
        return item

    def import_review(self, entity_title, rating, title, review_url, content, time):
        """
        Import one review: return 1: done / 2: skipped / None: failed
        """
        prefix = f"{self.user} |"
        url = self.guess_entity_url(entity_title, rating, time)
        if url is None:
            print(f"{prefix} fetching review {review_url}")
            try:
                h = DoubanDownloader(review_url).download().html()
                for u in h.xpath("//header[@class='main-hd']/a/@href"):
                    if ".douban.com/subject/" in u:
                        url = u
                if not url:
                    print(
                        f"{prefix} fetching error {review_url} unable to locate entity url"
                    )
                    return
            except Exception:
                print(f"{prefix} fetching review exception {review_url}")
                return
        item = self.get_item_by_url(url)
        if not item:
            print(f"{prefix} match/fetch {url} failed")
            return
        if Review.objects.filter(owner=self.user, item=item).exists():
            return 2
        content = re.sub(
            r'<span style="font-weight: bold;">([^<]+)</span>', r"<b>\1</b>", content
        )
        content = re.sub(r"(<img [^>]+>)", r"\1<br>", content)
        content = re.sub(
            r'<div class="image-caption">([^<]+)</div>', r"<br><i>\1</i><br>", content
        )
        content = md(content)
        content = re.sub(
            r"(?<=!\[\]\()([^)]+)(?=\))", lambda x: _fetch_remote_image(x[1]), content
        )
        params = {
            "owner": self.user,
            "created_time": time,
            "edited_time": time,
            "title": title,
            "body": content,
            "visibility": self.visibility,
            "item": item,
        }
        Review.objects.create(**params)
        return 1
