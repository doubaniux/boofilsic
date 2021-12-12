from django.shortcuts import reverse, redirect, render, get_object_or_404
from django.http import HttpResponseBadRequest, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import auth
from django.contrib.auth import authenticate
from django.core.paginator import Paginator
from django.utils.translation import gettext_lazy as _
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Count
from .models import User, Report, Preference
from .forms import ReportForm
from mastodon.auth import *
from mastodon.api import *
from mastodon import mastodon_request_included
from common.config import *
from common.models import MarkStatusEnum
from common.utils import PageLinksGenerator
from management.models import Announcement
from books.models import *
from movies.models import *
from music.models import *
from games.models import *
from books.forms import BookMarkStatusTranslator
from movies.forms import MovieMarkStatusTranslator
from music.forms import MusicMarkStatusTranslator
from games.forms import GameMarkStatusTranslator
from mastodon.models import MastodonApplication
from django.conf import settings
from urllib.parse import quote
from openpyxl import Workbook
from common.utils import GenerateDateUUIDMediaFilePath
from datetime import datetime
import os


def export_marks_task(user):
    user.preference.export_status['marks_pending'] = True
    user.preference.save()
    filename = GenerateDateUUIDMediaFilePath(None, 'f.xlsx', settings.MEDIA_ROOT + settings.EXPORT_FILE_PATH_ROOT)
    if not os.path.exists(os.path.dirname(filename)):
        os.makedirs(os.path.dirname(filename))
    heading = ['标题', '简介', '豆瓣评分', '链接', '创建时间', '我的评分', '标签', '评论', 'NeoDB链接', '其它ID']
    wb = Workbook()  # adding write_only=True will speed up but corrupt the xlsx and won't be importable
    for status, label in [('collect', '看过'), ('do', '在看'), ('wish', '想看')]:
        ws = wb.create_sheet(title=label)
        marks = MovieMark.objects.filter(owner=user, status=status).order_by("-edited_time")
        ws.append(heading)
        for mark in marks:
            movie = mark.movie
            title = movie.title
            summary = str(movie.year) + ' / ' + ','.join(movie.area) + ' / ' + ','.join(map(lambda x: str(MovieGenreTranslator[x]), movie.genre)) + ' / ' + ','.join(movie.director) + ' / ' + ','.join(movie.actor)
            tags = ','.join(list(set(map(lambda t: t['content'], movie.get_tags_manager().values()))))
            world_rating = (movie.rating / 2) if movie.rating else None
            timestamp = mark.edited_time.strftime('%Y-%m-%d %H:%M:%S')
            my_rating = (mark.rating / 2) if mark.rating else None
            text = mark.text
            source_url = movie.source_url
            url = settings.APP_WEBSITE + movie.get_absolute_url()
            line = [title, summary, world_rating, source_url, timestamp, my_rating, tags, text, url, movie.imdb_code]
            ws.append(line)

    for status, label in [('collect', '听过'), ('do', '在听'), ('wish', '想听')]:
        ws = wb.create_sheet(title=label)
        marks = AlbumMark.objects.filter(owner=user, status=status).order_by("-edited_time")
        ws.append(heading)
        for mark in marks:
            album = mark.album
            title = album.title
            summary = ','.join(album.artist) + ' / ' + (album.release_date.strftime('%Y') if album.release_date else '')
            tags = ','.join(list(set(map(lambda t: t['content'], album.get_tags_manager().values()))))
            world_rating = (album.rating / 2) if album.rating else None
            timestamp = mark.edited_time.strftime('%Y-%m-%d %H:%M:%S')
            my_rating = (mark.rating / 2) if mark.rating else None
            text = mark.text
            source_url = album.source_url
            url = settings.APP_WEBSITE + album.get_absolute_url()
            line = [title, summary, world_rating, source_url, timestamp, my_rating, tags, text, url, '']
            ws.append(line)

    for status, label in [('collect', '读过'), ('do', '在读'), ('wish', '想读')]:
        ws = wb.create_sheet(title=label)
        marks = BookMark.objects.filter(owner=user, status=status).order_by("-edited_time")
        ws.append(heading)
        for mark in marks:
            book = mark.book
            title = book.title
            summary = ','.join(book.author) + ' / ' + str(book.pub_year) + ' / ' + book.pub_house
            tags = ','.join(list(set(map(lambda t: t['content'], book.get_tags_manager().values()))))
            world_rating = (book.rating / 2) if book.rating else None
            timestamp = mark.edited_time.strftime('%Y-%m-%d %H:%M:%S')
            my_rating = (mark.rating / 2) if mark.rating else None
            text = mark.text
            source_url = book.source_url
            url = settings.APP_WEBSITE + book.get_absolute_url()
            line = [title, summary, world_rating, source_url, timestamp, my_rating, tags, text, url, book.isbn]
            ws.append(line)

    for status, label in [('collect', '玩过'), ('do', '在玩'), ('wish', '想玩')]:
        ws = wb.create_sheet(title=label)
        marks = GameMark.objects.filter(owner=user, status=status).order_by("-edited_time")
        ws.append(heading)
        for mark in marks:
            game = mark.game
            title = game.title
            summary = ','.join(game.genre) + ' / ' + ','.join(game.platform) + ' / ' + game.release_date.strftime('%Y-%m-%d')
            tags = ','.join(list(set(map(lambda t: t['content'], game.get_tags_manager().values()))))
            world_rating = (game.rating / 2) if game.rating else None
            timestamp = mark.edited_time.strftime('%Y-%m-%d %H:%M:%S')
            my_rating = (mark.rating / 2) if mark.rating else None
            text = mark.text
            source_url = game.source_url
            url = settings.APP_WEBSITE + game.get_absolute_url()
            line = [title, summary, world_rating, source_url, timestamp, my_rating, tags, text, url, '']
            ws.append(line)

    wb.save(filename=filename)
    user.preference.export_status['marks_pending'] = False
    user.preference.export_status['marks_file'] = filename
    user.preference.export_status['marks_date'] = datetime.now().strftime("%Y-%m-%d %H:%M")
    user.preference.save()
