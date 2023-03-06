from django.utils.translation import gettext_lazy as _
from .models import User
from mastodon.api import *
from common.config import *
from books.models import *
from movies.models import *
from music.models import *
from games.models import *
from django.conf import settings
from openpyxl import Workbook
from common.utils import GenerateDateUUIDMediaFilePath
from datetime import datetime
import os
from tqdm import tqdm
from datetime import timedelta


def refresh_user_mastodon_data_task(user, token=None):
    if token:
        user.mastodon_token = token
    if user.refresh_mastodon_data():
        user.save()
        print(f"{user} mastodon data refreshed")
    else:
        print(f"{user} mastodon data refresh failed")


def export_marks_task(user, file=None):
    user.preference.export_status["marks_pending"] = True
    user.preference.save(update_fields=["export_status"])
    filename = file or GenerateDateUUIDMediaFilePath(
        None, "f.xlsx", settings.MEDIA_ROOT + settings.EXPORT_FILE_PATH_ROOT
    )
    if not os.path.exists(os.path.dirname(filename)):
        os.makedirs(os.path.dirname(filename))
    heading = ["标题", "简介", "豆瓣评分", "链接", "创建时间", "我的评分", "标签", "评论", "NeoDB链接", "其它ID"]
    wb = (
        Workbook()
    )  # adding write_only=True will speed up but corrupt the xlsx and won't be importable
    for status, label in [("collect", "看过"), ("do", "在看"), ("wish", "想看")]:
        ws = wb.create_sheet(title=label)
        marks = MovieMark.objects.filter(owner=user, status=status).order_by(
            "-edited_time"
        )
        ws.append(heading)
        for mark in marks:
            movie = mark.movie
            title = movie.title
            summary = (
                str(movie.year)
                + " / "
                + ",".join(movie.area)
                + " / "
                + ",".join(map(lambda x: str(MovieGenreTranslator[x]), movie.genre))
                + " / "
                + ",".join(movie.director)
                + " / "
                + ",".join(movie.actor)
            )
            tags = ",".join(list(map(lambda m: m.content, mark.tags)))
            world_rating = (movie.rating / 2) if movie.rating else None
            timestamp = mark.edited_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = (mark.rating / 2) if mark.rating else None
            text = mark.text
            source_url = movie.source_url
            url = settings.APP_WEBSITE + movie.get_absolute_url()
            line = [
                title,
                summary,
                world_rating,
                source_url,
                timestamp,
                my_rating,
                tags,
                text,
                url,
                movie.imdb_code,
            ]
            ws.append(line)

    for status, label in [("collect", "听过"), ("do", "在听"), ("wish", "想听")]:
        ws = wb.create_sheet(title=label)
        marks = AlbumMark.objects.filter(owner=user, status=status).order_by(
            "-edited_time"
        )
        ws.append(heading)
        for mark in marks:
            album = mark.album
            title = album.title
            summary = (
                ",".join(album.artist)
                + " / "
                + (album.release_date.strftime("%Y") if album.release_date else "")
            )
            tags = ",".join(list(map(lambda m: m.content, mark.tags)))
            world_rating = (album.rating / 2) if album.rating else None
            timestamp = mark.edited_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = (mark.rating / 2) if mark.rating else None
            text = mark.text
            source_url = album.source_url
            url = settings.APP_WEBSITE + album.get_absolute_url()
            line = [
                title,
                summary,
                world_rating,
                source_url,
                timestamp,
                my_rating,
                tags,
                text,
                url,
                "",
            ]
            ws.append(line)

    for status, label in [("collect", "读过"), ("do", "在读"), ("wish", "想读")]:
        ws = wb.create_sheet(title=label)
        marks = BookMark.objects.filter(owner=user, status=status).order_by(
            "-edited_time"
        )
        ws.append(heading)
        for mark in marks:
            book = mark.book
            title = book.title
            summary = (
                ",".join(book.author)
                + " / "
                + str(book.pub_year)
                + " / "
                + book.pub_house
            )
            tags = ",".join(list(map(lambda m: m.content, mark.tags)))
            world_rating = (book.rating / 2) if book.rating else None
            timestamp = mark.edited_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = (mark.rating / 2) if mark.rating else None
            text = mark.text
            source_url = book.source_url
            url = settings.APP_WEBSITE + book.get_absolute_url()
            line = [
                title,
                summary,
                world_rating,
                source_url,
                timestamp,
                my_rating,
                tags,
                text,
                url,
                book.isbn,
            ]
            ws.append(line)

    for status, label in [("collect", "玩过"), ("do", "在玩"), ("wish", "想玩")]:
        ws = wb.create_sheet(title=label)
        marks = GameMark.objects.filter(owner=user, status=status).order_by(
            "-edited_time"
        )
        ws.append(heading)
        for mark in marks:
            game = mark.game
            title = game.title
            summary = (
                ",".join(game.genre)
                + " / "
                + ",".join(game.platform)
                + " / "
                + (game.release_date.strftime("%Y-%m-%d") if game.release_date else "")
            )
            tags = ",".join(list(map(lambda m: m.content, mark.tags)))
            world_rating = (game.rating / 2) if game.rating else None
            timestamp = mark.edited_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = (mark.rating / 2) if mark.rating else None
            text = mark.text
            source_url = game.source_url
            url = settings.APP_WEBSITE + game.get_absolute_url()
            line = [
                title,
                summary,
                world_rating,
                source_url,
                timestamp,
                my_rating,
                tags,
                text,
                url,
                "",
            ]
            ws.append(line)

    review_heading = [
        "标题",
        "评论对象",
        "链接",
        "创建时间",
        "我的评分",
        "类型",
        "内容",
        "评论对象原始链接",
        "评论对象NeoDB链接",
    ]
    for ReviewModel, label in [
        (MovieReview, "影评"),
        (BookReview, "书评"),
        (AlbumReview, "乐评"),
        (GameReview, "游戏评论"),
    ]:
        ws = wb.create_sheet(title=label)
        reviews = ReviewModel.objects.filter(owner=user).order_by("-edited_time")
        ws.append(review_heading)
        for review in reviews:
            title = review.title
            target = "《" + review.item.title + "》"
            url = review.absolute_url
            timestamp = review.edited_time.strftime("%Y-%m-%d %H:%M:%S")
            my_rating = None  # (mark.rating / 2) if mark.rating else None
            content = review.content
            target_source_url = review.item.source_url
            target_url = review.item.absolute_url
            line = [
                title,
                target,
                url,
                timestamp,
                my_rating,
                label,
                content,
                target_source_url,
                target_url,
            ]
            ws.append(line)

    wb.save(filename=filename)
    user.preference.export_status["marks_pending"] = False
    user.preference.export_status["marks_file"] = filename
    user.preference.export_status["marks_date"] = datetime.now().strftime(
        "%Y-%m-%d %H:%M"
    )
    user.preference.save(update_fields=["export_status"])


def refresh_mastodon_relationships_task():
    count = 0
    for user in tqdm(
        User.objects.filter(
            mastodon_last_refresh__lt=timezone.now()
            - timedelta(hours=settings.MASTODON_RELATIONSHIP_CACHE_LIFETIME),
            is_active=True,
        )
    ):
        if user.mastodon_token or user.mastodon_refresh_token:
            tqdm.write(f"Refreshing {user}")
            if user.refresh_mastodon_data():
                tqdm.write(f"Refreshed {user}")
                count += 1
            else:
                tqdm.write(f"Refresh failed for {user}")
            user.save()
        else:
            tqdm.write(f"Missing token for {user}")

    print(f"{count} users updated")
